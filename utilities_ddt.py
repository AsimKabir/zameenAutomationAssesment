import openpyxl as xl
# from openpyxl import xl
from openpyxl.worksheet.worksheet import Worksheet
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import requests 
import config as config

def getRowCount(file, sheetName):
    
    workbook = xl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)
    

def getColumnCount(file, sheetName):
    
    workbook = xl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)
    

def readData(file, sheetName, rowNumber, columnNumber):
    
    workbook = xl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.cell(row = rowNumber, column = columnNumber).value)
    

def writeData(file, sheetName, rowNumber, columnNumber, data):
    workbook = xl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row = rowNumber, column = columnNumber).value = data
    workbook.save(file)


def getAllLinks(driver):
    f=open('./textFiles/linkssitemap.txt','w')
    elems = driver.find_elements_by_tag_name('a')
    href = ""
    for elem in elems:
        # href = elem.get_attribute('href')
        href = elem.text
        if href is not None:
            # if href != "javascript:void(0)" and href != "tel:04238900939" and href != "mailto:care@oladoc.com" and href != "tel:0518151800" and href!="tel:02138140600" and href!="javascript:void(0);":
                f.write(href+'\n')
                print(href)
    f.close()
    # return href


def getAllMetaTags():
    f=open('./textFiles/Meta.txt','w')
    f.write("URL::" + driver.current_url + "\n-----------------------------------------------------\n\n")
    f.write("TITLE of Page::" + driver.title + "\n-----------------------------------------------------\n\n")
    elems = driver.find_elements_by_tag_name('meta')
    tagname=""
    content = ""
    proper = ""
    name = ""
    for elem in elems:
        tagname = elem.tag_name
        content = elem.get_attribute('content')
        # pro = elem.get_attribute('property')
        name = elem.get_attribute('name')
        if content is not None:
            f.write("tagname  --->"+tagname+'\n')
            # f.write("Property  --->"+pro+'\n')
            f.write("Name     --->"+name+'\n')
            f.write("content  --->"+content+'\n\n\n')
            print(content)
    f.close()
    # return href


def saveListInToTextFile ():
    l1=['hi','hello','welcome']
    f=open('./textFiles/f1.txt','w')
    for ele in l1:
        f.write(ele+'\n')
    f.close()

def checkStatusOfURL(urlList):

    for x in urlList:
        print(x)
        response = requests.get(x) 
  
        # print response 
        # print(response) 
        # print url 
        # print(response.url) 
        codee = response.status_code
        print(codee)
        
        # pass

def fileIntoList():
    text_file = open("./textFiles/links.txt", "r")
    lines = text_file.readlines()
    # for x in lines:
    #     print (x)
        
    # print (len(lines))
    text_file.close()

    return lines




url = config.KEYS["BASE_URL"]
driver = config.driver

# Explicit wait
# try:
#     element = WebDriverWait(driver, 10).until(
#         EC.presence_of_element_located((By.ID, "myDynamicElement"))
#     )
# finally:
#     driver.quit()

# implicit wait
driver.implicitly_wait(30) # seconds


driver.get(url)

driver.maximize_window()
getAllLinks(driver)
print ("SSS")
# getAllMetaTags()


# print("\n\n\n\n\n<<<<------>>>>>\n\n\n\n\n")
# print()

# saveListInToTextFile()

links = fileIntoList()
# checkStatusOfURL(links)

driver.close()

#it is end of the page