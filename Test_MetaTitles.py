# from os import pathsep
import unittest as ut
import logging as lg
import openpyxl
# import utilities_ddt as utility
# import config as config



class test_OladocMetaTitles(ut.TestCase):
    
    @classmethod
    def setUpClass(self):
        super(test_OladocMetaTitles, self).setUpClass()
        
        lg.basicConfig(filename="C://Users/Asim Kabir/Desktop/dev/SeleniumWebDriver Practice/master/oladoc/logs/metaTitles_logs.txt", format='%(asctime)s: %(levelname)s:---> %(message)s', datefmt='%m/%d/%y %I:%M:%S:%p', level=lg.DEBUG)
        lg.debug("Class setup is done")
        lg.info("it is info")
        lg.warning("it is warning")
        lg.error("it is error")
        lg.critical("It is critical")


        path="C://Users/Asim Kabir/Desktop/dev/SeleniumWebDriver Practice/master/oladoc/dataFiles/oladoc_Links.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.get_sheet_by_name("specialityListing")  # sheet = workbook.active
        
        rows = sheet.max_row
        print(rows)
        columns = sheet.max_column
        print(columns)

        
    

    def test_HomePageMetaTitles(self):
        title = "Hello"
        self.assertEqual("Helloo", title, "FAILED:: Page Title is wrong")
        
    def test_SpecialityListingMetaTitles(self):
        title = "Hello"
        self.assertEqual("Hello", title, "FAILED:: Title is wrong")
        
    def test_TreatmentListingMetaTitles(self):
        title = "Hello"
        self.assertEqual("Hello", title, "FAILED:: Title is wrong")
        
    def test_ConditionListingMetaTitles(self):
        title = "Hello"
        self.assertEqual("Helslo", title, "FAILED:: As Title is wrong")
        
    def  test_BookableDoctorProfileMetaTitles(self):
        title = "Hello"
        self.assertEqual("Hello", title, "FAILED:: Title is wrong")
        
    def  test_NonBookableDoctorProfileMetaTitles(self):
        title = "Hello"
        self.assertEqual("Hello", title, "FAILED:: Title is wrong")
        
if __name__== "__main__":
    ut.main()